import streamlit as st
import pandas as pd
import datetime
import io
from src.ui.components import setup_page, success_box, display_footer

# Konfiguracja strony
setup_page()

with st.sidebar:
    st.logo(
        'img/logo.svg',
        size='large'
    )

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

# Dodanie zmiennej do przechowywania profili
if "file_profiles" not in st.session_state:
    st.session_state.file_profiles = {}

st.markdown("""<div style='text-align: center; padding: 1rem; border-radius: 5px; margin-bottom: 2rem;'><h1>Kalkulator PTPIRE</h1><p>Narzędzie do analizy plików pomiarowych</p></div>""", unsafe_allow_html=True)

st.subheader("Wgraj pliki PTPIRE")

uploaded_files = st.file_uploader("Wybierz pliki tekstowe", type="dat", accept_multiple_files=True, key=f"uploader_{st.session_state.uploader_key}")

def update_key():
    st.session_state.uploader_key += 1
    # Reset profili przy czyszczeniu danych
    st.session_state.file_profiles = {}

def group_files_by_name(files):
    grouped_files = {}
    for file in files:
        name = file.name.split("_2")[0]
        if name not in grouped_files:
            grouped_files[name] = []
        grouped_files[name].append(file)
    return grouped_files

def get_profile_from_file(file):
    content = file.read().decode("utf-8")
    lines = content.split("\n")
    profile = []
    for line in lines[6:-2]:
        try:
            profile.append(float(line.strip()[:-2]))
        except ValueError:
            continue
    return profile

def get_profile_date(file):
    content = file.read().decode("utf-8")
    lines = content.split("\n")
    try:
        profile_date = lines[4].strip()
    except (IndexError, ValueError):
        profile_date = None
    return profile_date

def convert_to_hourly(profile):
    hourly_profile = []
    for i in range(0, len(profile), 4):
        hourly_sum = sum(profile[i:i+4])
        hourly_profile.append(hourly_sum)
    return hourly_profile

# --- ROZDZIELENIE PLIKÓW NA DWIE GRUPY ---
cp_files = [f for f in uploaded_files if '_CP_' in f.name and '_CPP_' not in f.name]
co_files = [f for f in uploaded_files if '_CO_' in f.name and '_CPP_' not in f.name]

def process_files(files):
    grouped_files = group_files_by_name(files)
    data = {"Lp": [], "Nazwa": [], "Data": [], "Suma": [], "Pliki": [], "Profil": []}
    lp = 0
    for name, files in grouped_files.items():
        total_sum = 0
        counter = 0
        profile = []
        lp += 1
        for uploaded_file in files:
            uploaded_file.seek(0)
            file_profile = get_profile_from_file(uploaded_file)
            uploaded_file.seek(0)
            profile_date = get_profile_date(uploaded_file)
            if len(file_profile) > 25:
                file_profile = convert_to_hourly(file_profile)
            file_key = f"{name}_{counter+1}"
            st.session_state.file_profiles[file_key] = {
                "profile": file_profile,
                "date": profile_date
            }
            total_sum += sum(file_profile)
            profile.extend(file_profile)
            counter += 1
        data["Lp"].append(lp)
        data["Nazwa"].append(name)
        data["Data"].append(profile_date)
        data["Suma"].append(total_sum)
        data["Pliki"].append(counter)
        data["Profil"].append(profile)
    return pd.DataFrame(data)

# --- PRZETWARZANIE DANYCH DLA OBU GRUP ---
df_cp = process_files(cp_files)
df_co = process_files(co_files)

# --- WYŚWIETLANIE TABEL ---
if not df_cp.empty:
    st.markdown("<h3 style='text-align:center;'>Energia czynna pobrana</h3>", unsafe_allow_html=True)
    st.dataframe(
        df_cp.style.format({"Suma": "{:.3f}"}),
        column_config={
            "Profil": st.column_config.BarChartColumn(
                "Profil",
                y_min=0,
                y_max=df_cp["Profil"].apply(lambda x: max(x)).max() if not df_cp.empty else 0
            )
        },
        use_container_width=True,
        hide_index=True
    )

if not df_co.empty:
    st.markdown("<h3 style='text-align:center;'>Energia czynna oddana</h3>", unsafe_allow_html=True)
    st.dataframe(
        df_co.style.format({"Suma": "{:.3f}"}),
        column_config={
            "Profil": st.column_config.BarChartColumn(
                "Profil",
                y_min=0,
                y_max=df_co["Profil"].apply(lambda x: max(x)).max() if not df_co.empty else 0
            )
        },
        use_container_width=True,
        hide_index=True
    )

# --- EKSPORT DO EXCELA DLA ENERGII CZYNNEJ POBRANEJ (CP) ---
if not df_cp.empty:
    df_to_excel_cp = df_cp.copy().drop('Profil', axis=1)
    output_cp = io.BytesIO()
    with pd.ExcelWriter(output_cp, engine='openpyxl') as writer:
        writer.sheets['Sheet1'] = writer.book.create_sheet('Sheet1')
        cell = writer.sheets['Sheet1'].cell(row=1, column=2, value="Raport ilości plików PTPiREE - Energia czynna pobrana")
        from openpyxl.styles import Font, numbers
        cell.font = Font(size=14, bold=True)
        df_to_excel_cp.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
        ws = writer.sheets['Sheet1']
        suma_col = None
        for idx, col in enumerate(df_to_excel_cp.columns, 1):
            if col == 'Suma':
                suma_col = idx
                break
        if suma_col:
            for row in range(4, len(df_to_excel_cp) + 4):
                cell = ws.cell(row=row, column=suma_col)
                cell.number_format = '0.000'
            last_row = len(df_to_excel_cp) + 5
            ws.cell(row=last_row, column=2, value="Ilość plików sumarycznie")
            total_files_cell = ws.cell(row=last_row, column=3, value=int(df_to_excel_cp['Pliki'].sum()))
            total_files_cell.number_format = '0'
            ws.cell(row=last_row + 1, column=2, value="Całkowity wolumen")
            total_sum_cell = ws.cell(row=last_row + 1, column=3, value=float(df_to_excel_cp['Suma'].sum()))
            total_sum_cell.number_format = '0.000'
            ws.cell(row=last_row + 2, column=2, value="Data wykonania raportu")
            date_cell = ws.cell(row=last_row + 2, column=3, value=datetime.datetime.now())
            date_cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    now = datetime.datetime.now()
    output_cp.seek(0)
    st.download_button(
        "Pobierz plik xlsx (Energia czynna pobrana)",
        output_cp,
        f"ptpire_CP_{now.strftime('%Y%m%d_%H%M%S')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key='download-excel-cp'
    )

# --- EKSPORT DO EXCELA DLA ENERGII CZYNNEJ ODDANEJ (CO) ---
if not df_co.empty:
    df_to_excel_co = df_co.copy().drop('Profil', axis=1)
    output_co = io.BytesIO()
    with pd.ExcelWriter(output_co, engine='openpyxl') as writer:
        writer.sheets['Sheet1'] = writer.book.create_sheet('Sheet1')
        cell = writer.sheets['Sheet1'].cell(row=1, column=2, value="Raport ilości plików PTPiREE - Energia czynna oddana")
        from openpyxl.styles import Font, numbers
        cell.font = Font(size=14, bold=True)
        df_to_excel_co.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
        ws = writer.sheets['Sheet1']
        suma_col = None
        for idx, col in enumerate(df_to_excel_co.columns, 1):
            if col == 'Suma':
                suma_col = idx
                break
        if suma_col:
            for row in range(4, len(df_to_excel_co) + 4):
                cell = ws.cell(row=row, column=suma_col)
                cell.number_format = '0.000'
            last_row = len(df_to_excel_co) + 5
            ws.cell(row=last_row, column=2, value="Ilość plików sumarycznie")
            total_files_cell = ws.cell(row=last_row, column=3, value=int(df_to_excel_co['Pliki'].sum()))
            total_files_cell.number_format = '0'
            ws.cell(row=last_row + 1, column=2, value="Całkowity wolumen")
            total_sum_cell = ws.cell(row=last_row + 1, column=3, value=float(df_to_excel_co['Suma'].sum()))
            total_sum_cell.number_format = '0.000'
            ws.cell(row=last_row + 2, column=2, value="Data wykonania raportu")
            date_cell = ws.cell(row=last_row + 2, column=3, value=datetime.datetime.now())
            date_cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    now = datetime.datetime.now()
    output_co.seek(0)
    st.download_button(
        "Pobierz plik xlsx (Energia czynna oddana)",
        output_co,
        f"ptpire_CO_{now.strftime('%Y%m%d_%H%M%S')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key='download-excel-co'
    )

# Dodawanie przycisku do eksportu profili
if st.session_state.file_profiles:
    # Tworzenie danych do eksportu profili
    profiles_output = io.BytesIO()
    
    with pd.ExcelWriter(profiles_output, engine='openpyxl') as writer:
        # Przygotuj dane profili do zapisu
        for group_name, files in group_files_by_name(uploaded_files).items():
            # Pomiń grupy i pliki z '_CPP_'
            if '_CPP_' in group_name:
                continue
            all_profiles_data = []
            max_length = 0
            for file in files:
                if '_CPP_' in file.name:
                    continue
                file_key = f"{group_name}_{files.index(file)+1}"
                if file_key not in st.session_state.file_profiles:
                    continue
                profile = st.session_state.file_profiles[file_key]["profile"]
                max_length = max(max_length, len(profile))
            for file in files:
                if '_CPP_' in file.name:
                    continue
                file_key = f"{group_name}_{files.index(file)+1}"
                if file_key not in st.session_state.file_profiles:
                    continue
                profile = st.session_state.file_profiles[file_key]["profile"]
                profile_date = st.session_state.file_profiles[file_key]["date"]
                padded_profile = profile + [None] * (max_length - len(profile))
                row_data = {"Nazwa pliku": file_key, "Data profilu": profile_date}
                for i, value in enumerate(padded_profile):
                    row_data[f"{i+1}"] = value
                all_profiles_data.append(row_data)
            if not all_profiles_data:
                continue
            profiles_df = pd.DataFrame(all_profiles_data)
            # Sortowanie po kolumnie "Data profilu"
            # Konwersja kolumny na datetime
            profiles_df["Data profilu"] = pd.to_datetime(profiles_df["Data profilu"], format="%d-%m-%Y", errors="coerce")
            profiles_df = profiles_df.sort_values(by="Data profilu", ascending=True, na_position='last')
            # Zamiana na tekstowy format daty
            profiles_df["Data profilu"] = profiles_df["Data profilu"].dt.strftime("%d-%m-%Y")

            profiles_df.to_excel(writer, sheet_name=group_name, index=False, startrow=2)
            ws = writer.sheets[group_name]
            ws.cell(row=1, column=1, value=f"Profile plików PTPiREE - {group_name}")
            for row in range(4, len(profiles_df) + 4):
                for col in range(2, max_length + 3):
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = '0.000'
            ws.cell(row=len(profiles_df) + 5, column=1, value="Data wygenerowania")
            date_cell = ws.cell(row=len(profiles_df) + 5, column=2, value=datetime.datetime.now())
            date_cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    
    now = datetime.datetime.now()
    profiles_output.seek(0)
    
    st.download_button(
        "Pobierz profile jako xlsx",
        profiles_output,
        f"ptpire_profiles_{now.strftime('%Y%m%d_%H%M%S')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key='download-profiles'
    )
    
    # Dodaj przycisk do czyszczenia wszystkich kontrolek i stanu
    if st.button("Wyczyść wszystkie dane", type="primary"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        # Zwiększ klucz uploadu, wyczyść file_profiles i uploaded_files
        st.session_state.uploader_key = st.session_state.get('uploader_key', 0) + 1
        st.session_state.file_profiles = {}
        st.session_state[f'uploader_{st.session_state.uploader_key}'] = []
        st.rerun()
# Wyświetl stopkę
display_footer()




